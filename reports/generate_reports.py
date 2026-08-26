#!/usr/bin/env python3
"""
generate_reports.py — builds the daily Excel report + HTML dashboard from a
recompute_top_of_book.csv export (produced by refresh.sh / bigquery/recompute_top_of_book.sql).

Usage:
    python3 reports/generate_reports.py <path-to-recompute_top_of_book.csv> [scan-date YYYY-MM-DD]

Outputs (written next to the CSV, unless --outdir is given):
    jita_denni_report_<date>.xlsx   (formulas NOT yet recalculated — see note below)
    jita_dashboard.html             (self-contained, no external deps except Google Fonts)

Requires: openpyxl (pip install openpyxl --break-system-packages if missing)

IMPORTANT — two manual steps after running this, which only Claude can do (not this script):
  1. Run the xlsx skill's `scripts/recalc.py` on the .xlsx output. openpyxl writes formulas as
     text with no cached values — most viewers (including a quick look at the file) will show
     blank cells in the "Kumulativní cena" column until LibreOffice recalculates it once.
  2. Publish/update jita_dashboard.html via the Artifact tool. If updating the existing dashboard
     Matej already has, pass its existing artifact URL so it republishes to the same link instead
     of creating a new one — see docs/eve-jita-scanner-ops.md for that URL.

Sizing/ranking methodology (see docs/eve-jita-scanner-ops.md for the full history/rationale):
  - suggested_units = round(0.15 * avg_daily_volume_14d)  — 15% of daily turnover, NO capital cap
  - suggested_cost  = suggested_units * buy_price
  - suggested_profit = suggested_units * profit_per_unit
  - ranked descending by suggested_profit (not by margin_pct or profit_per_unit alone)
  - fees already baked into profit_per_unit / margin_pct columns of the input CSV (broker
    1.382%, sales tax 3.375%) — don't re-apply them here.
"""
import csv
import json
import sys
from pathlib import Path


def load_rows(csv_path):
    rows = []
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            if not r.get("type_id"):
                continue
            buy = float(r["buy_price"])
            sell = float(r["sell_price"])
            margin = float(r["margin_pct"])
            buy_orders = int(r["buy_orders"])
            sell_orders = int(r["sell_orders"])
            vol = float(r["avg_daily_volume_14d"])
            ppu = float(r["profit_per_unit"])
            units = round(0.15 * vol)
            rows.append({
                "name": r["item_name"],
                "buy": round(buy, 2),
                "sell": round(sell, 2),
                "margin": round(margin, 3),
                "buyOrders": buy_orders,
                "sellOrders": sell_orders,
                "volume": round(vol, 1),
                "ppu": round(ppu, 2),
                "units": int(units),
                "cost": round(units * buy),
                "profit": round(units * ppu),
            })
    rows.sort(key=lambda x: -x["profit"])
    for i, r in enumerate(rows, 1):
        r["rank"] = i
    return rows


# ---------------------------------------------------------------------------
# Dashboard (self-contained HTML, hand-rolled SVG charts — no Chart.js/CDN,
# so it also works if you ever publish it via the Artifact tool, which blocks
# external script hosts other than Google Fonts).
# ---------------------------------------------------------------------------
DASHBOARD_TEMPLATE = r"""<!doctype html>
<title>Jita Obchodní Radar</title>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@400;500;600;700;800&family=IBM+Plex+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  :root {
    color-scheme: light;
    --bg: #f4f3ef;
    --surface: #ffffff;
    --surface-2: #ecebe5;
    --border: #dedcd3;
    --text-primary: #17181a;
    --text-secondary: #5b5c54;
    --text-muted: #8a8b80;
    --accent: #a86a14;
    --accent-soft: #f1e3cc;
    --slot-1: #2a78d6;
    --slot-2: #eb6834;
    --slot-3: #1baf7a;
    --good: #0ca30c;
    --grid: #e4e2d8;
    --shadow: 0 1px 2px rgba(30,28,20,0.06), 0 8px 24px rgba(30,28,20,0.05);
    --radius: 10px;
  }
  @media (prefers-color-scheme: dark) {
    :root:not([data-theme="light"]) {
      color-scheme: dark;
      --bg: #0e1013;
      --surface: #171a1e;
      --surface-2: #1e2126;
      --border: #2b2e34;
      --text-primary: #f2f1ea;
      --text-secondary: #a3a59b;
      --text-muted: #74766e;
      --accent: #dba54a;
      --accent-soft: #3a2f1a;
      --slot-1: #3987e5;
      --slot-2: #d95926;
      --slot-3: #199e70;
      --good: #3fca3f;
      --grid: #24272c;
      --shadow: 0 1px 2px rgba(0,0,0,0.3), 0 8px 24px rgba(0,0,0,0.35);
    }
  }
  :root[data-theme="dark"] {
    color-scheme: dark;
    --bg: #0e1013;
    --surface: #171a1e;
    --surface-2: #1e2126;
    --border: #2b2e34;
    --text-primary: #f2f1ea;
    --text-secondary: #a3a59b;
    --text-muted: #74766e;
    --accent: #dba54a;
    --accent-soft: #3a2f1a;
    --slot-1: #3987e5;
    --slot-2: #d95926;
    --slot-3: #199e70;
    --good: #3fca3f;
    --grid: #24272c;
    --shadow: 0 1px 2px rgba(0,0,0,0.3), 0 8px 24px rgba(0,0,0,0.35);
  }

  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; }
  body {
    background: var(--bg);
    color: var(--text-primary);
    font-family: "Sora", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
  }
  .mono { font-family: "IBM Plex Mono", ui-monospace, SFMono-Regular, monospace; }
  .num { font-variant-numeric: tabular-nums; }

  .wrap {
    max-width: 1320px;
    margin: 0 auto;
    padding: 28px 24px 64px;
    display: flex;
    flex-direction: column;
    gap: 20px;
  }

  header.top {
    display: flex;
    justify-content: space-between;
    align-items: flex-end;
    gap: 16px;
    flex-wrap: wrap;
    border-bottom: 1px solid var(--border);
    padding-bottom: 18px;
  }
  .eyebrow {
    font-family: "IBM Plex Mono", monospace;
    font-size: 11.5px;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--accent);
    margin: 0 0 6px;
  }
  h1 {
    font-size: 26px;
    font-weight: 800;
    letter-spacing: -0.01em;
    margin: 0 0 6px;
    text-wrap: balance;
  }
  header.top p {
    margin: 0;
    color: var(--text-secondary);
    font-size: 13.5px;
    max-width: 60ch;
  }
  .top-meta {
    text-align: right;
    font-family: "IBM Plex Mono", monospace;
    font-size: 12px;
    color: var(--text-muted);
    line-height: 1.7;
  }
  .top-meta strong { color: var(--text-secondary); font-weight: 500; }

  .kpi-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; }
  .kpi-card { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 16px 18px; box-shadow: var(--shadow); }
  .kpi-label { font-size: 11.5px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted); margin-bottom: 8px; }
  .kpi-value { font-family: "IBM Plex Mono", monospace; font-size: 25px; font-weight: 600; color: var(--text-primary); letter-spacing: -0.01em; }
  .kpi-sub { margin-top: 5px; font-size: 12px; color: var(--text-secondary); }
  .kpi-card.accent .kpi-value { color: var(--accent); }

  .filters { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); box-shadow: var(--shadow); padding: 14px 18px; display: flex; flex-wrap: wrap; align-items: end; gap: 18px; }
  .filter-group { display: flex; flex-direction: column; gap: 6px; }
  .filter-group label { font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; color: var(--text-muted); }
  .filter-group input[type="text"], .filter-group input[type="number"] { font-family: "IBM Plex Mono", monospace; font-size: 13px; padding: 7px 10px; border: 1px solid var(--border); border-radius: 6px; background: var(--surface-2); color: var(--text-primary); width: 130px; }
  .filter-group input[type="text"] { width: 200px; }
  .filter-group input:focus { outline: 2px solid var(--accent); outline-offset: 1px; }
  .chip-row { display: flex; gap: 6px; }
  .chip { font-size: 12px; padding: 6px 12px; border-radius: 999px; border: 1px solid var(--border); background: var(--surface-2); color: var(--text-secondary); cursor: pointer; user-select: none; display: flex; align-items: center; gap: 6px; }
  .chip .dot { width: 8px; height: 8px; border-radius: 50%; }
  .chip[data-active="true"] { background: var(--accent-soft); border-color: var(--accent); color: var(--text-primary); }
  .reset-btn { margin-left: auto; font-family: "IBM Plex Mono", monospace; font-size: 12px; background: transparent; border: 1px solid var(--border); border-radius: 6px; padding: 8px 12px; color: var(--text-secondary); cursor: pointer; }
  .reset-btn:hover { border-color: var(--accent); color: var(--accent); }
  .filter-count { font-family: "IBM Plex Mono", monospace; font-size: 12px; color: var(--text-muted); align-self: center; }

  .chart-row { display: grid; grid-template-columns: 1.3fr 1fr; gap: 12px; }
  .panel { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); box-shadow: var(--shadow); padding: 18px 20px 14px; }
  .panel h2 { font-size: 14px; font-weight: 600; margin: 0; }
  .panel .panel-sub { font-size: 12px; color: var(--text-muted); margin: 3px 0 12px; }
  svg text { font-family: "IBM Plex Mono", monospace; fill: var(--text-secondary); }
  .axis-line { stroke: var(--grid); stroke-width: 1; }
  .gridline { stroke: var(--grid); stroke-width: 1; }
  .bar-label { fill: var(--text-secondary); font-size: 10.5px; }
  .bar-value { fill: var(--text-primary); font-size: 11px; font-weight: 500; }

  .legend { display: flex; gap: 16px; margin-top: 10px; flex-wrap: wrap; }
  .legend-item { display: flex; align-items: center; gap: 6px; font-size: 11.5px; color: var(--text-secondary); }
  .legend-swatch { width: 9px; height: 9px; border-radius: 50%; }

  .viz-tooltip { position: fixed; pointer-events: none; background: var(--surface); border: 1px solid var(--border); border-radius: 8px; box-shadow: var(--shadow); padding: 9px 12px; font-size: 12px; color: var(--text-primary); z-index: 50; opacity: 0; transition: opacity 0.08s ease; max-width: 240px; }
  .viz-tooltip.show { opacity: 1; }
  .viz-tooltip .tt-title { font-weight: 600; margin-bottom: 4px; }
  .viz-tooltip .tt-row { display: flex; justify-content: space-between; gap: 14px; color: var(--text-secondary); font-family: "IBM Plex Mono", monospace; font-size: 11.5px; }
  .viz-tooltip .tt-row span:last-child { color: var(--text-primary); }

  .table-panel { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); box-shadow: var(--shadow); overflow: hidden; }
  .table-head-bar { display: flex; justify-content: space-between; align-items: center; padding: 16px 20px 10px; }
  .table-scroll { overflow: auto; max-height: 560px; border-top: 1px solid var(--border); }
  table.data-table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  .data-table thead th { position: sticky; top: 0; background: var(--surface-2); text-align: right; padding: 9px 14px; font-family: "IBM Plex Mono", monospace; font-size: 10.5px; text-transform: uppercase; letter-spacing: 0.04em; color: var(--text-muted); cursor: pointer; white-space: nowrap; border-bottom: 1px solid var(--border); user-select: none; }
  .data-table thead th:first-child, .data-table thead th:nth-child(2) { text-align: left; }
  .data-table thead th:hover { color: var(--accent); }
  .data-table thead th.sorted { color: var(--accent); }
  .data-table tbody td { padding: 8px 14px; text-align: right; font-family: "IBM Plex Mono", monospace; border-bottom: 1px solid var(--grid); white-space: nowrap; }
  .data-table tbody td:first-child, .data-table tbody td:nth-child(2) { text-align: left; font-family: "Sora", sans-serif; }
  .data-table tbody tr:hover td { background: var(--surface-2); }
  .rank-cell { color: var(--text-muted); }
  .tier-dot { display: inline-block; width: 7px; height: 7px; border-radius: 50%; margin-right: 7px; }
  .empty-row td { text-align: center; color: var(--text-muted); padding: 32px; font-family: "Sora", sans-serif; }

  footer.notes { font-size: 11.5px; color: var(--text-muted); border-top: 1px solid var(--border); padding-top: 14px; line-height: 1.7; }
  footer.notes strong { color: var(--text-secondary); }

  @media (max-width: 980px) {
    .kpi-row { grid-template-columns: repeat(2, 1fr); }
    .chart-row { grid-template-columns: 1fr; }
    .top-meta { text-align: left; }
  }
  @media (max-width: 640px) { .kpi-row { grid-template-columns: 1fr; } }
</style>

<div class="wrap">

  <header class="top">
    <div>
      <p class="eyebrow">EVE ONLINE · THE FORGE / JITA 4-4</p>
      <h1>Jita Obchodní Radar</h1>
      <p>Kandidáti na station trading seřazení podle dosažitelného zisku z pozice. Velikost pozice = 15 % průměrného denního obratu (14 dní) — bez kapitálového stropu.</p>
    </div>
    <div class="top-meta">
      Zdroj: eve-jita-poller (Cloud Run) · ceny top-of-book<br>
      Sken: <strong>__SCAN_DATE__</strong><br>
      Filtr: marže 0,1–100 %, min. 3 buy/sell orderů, min. 50 ks/den
    </div>
  </header>

  <section class="kpi-row" id="kpi-row"></section>

  <section class="filters">
    <div class="filter-group">
      <label for="f-search">Hledat položku</label>
      <input type="text" id="f-search" placeholder="např. Chromium">
    </div>
    <div class="filter-group">
      <label for="f-margin">Min. marže (%)</label>
      <input type="number" id="f-margin" value="0" min="0" max="100" step="1">
    </div>
    <div class="filter-group">
      <label for="f-volume">Min. obrat / den</label>
      <input type="number" id="f-volume" value="0" min="0" step="100">
    </div>
    <div class="filter-group">
      <label>Cenová hladina</label>
      <div class="chip-row" id="tier-chips">
        <div class="chip" data-tier="low" data-active="true"><span class="dot" style="background:var(--slot-1)"></span>Nízká (&lt;10k)</div>
        <div class="chip" data-tier="mid" data-active="true"><span class="dot" style="background:var(--slot-2)"></span>Střední (10k–1M)</div>
        <div class="chip" data-tier="high" data-active="true"><span class="dot" style="background:var(--slot-3)"></span>Vysoká (&gt;1M)</div>
      </div>
    </div>
    <div class="filter-count" id="filter-count"></div>
    <button class="reset-btn" id="reset-btn">Reset filtrů</button>
  </section>

  <section class="chart-row">
    <div class="panel">
      <h2>Top položky podle zisku z pozice</h2>
      <p class="panel-sub">Zisk pozice = doporučené množství × zisk/kus, filtrovaná sada</p>
      <svg id="bar-chart" width="100%" viewBox="0 0 720 420" preserveAspectRatio="xMinYMin meet"></svg>
    </div>
    <div class="panel">
      <h2>Marže × obrat × cenová hladina</h2>
      <p class="panel-sub">Bublina = zisk pozice · osa Y logaritmická</p>
      <svg id="scatter-chart" width="100%" viewBox="0 0 520 420" preserveAspectRatio="xMinYMin meet"></svg>
      <div class="legend">
        <div class="legend-item"><span class="legend-swatch" style="background:var(--slot-1)"></span>Nízká cena (&lt;10k ISK)</div>
        <div class="legend-item"><span class="legend-swatch" style="background:var(--slot-2)"></span>Střední (10k–1M ISK)</div>
        <div class="legend-item"><span class="legend-swatch" style="background:var(--slot-3)"></span>Vysoká (&gt;1M ISK)</div>
      </div>
    </div>
  </section>

  <section class="table-panel">
    <div class="table-head-bar">
      <h2 style="margin:0;font-size:14px;">Kompletní seznam kandidátů</h2>
      <span class="filter-count" id="table-count"></span>
    </div>
    <div class="table-scroll">
      <table class="data-table" id="data-table">
        <thead>
          <tr>
            <th data-field="rank" style="text-align:left">#</th>
            <th data-field="name" style="text-align:left">Položka</th>
            <th data-field="buy">Nákup (ISK)</th>
            <th data-field="sell">Prodej (ISK)</th>
            <th data-field="margin">Marže</th>
            <th data-field="ppu">Zisk/ks</th>
            <th data-field="volume">Obrat/den</th>
            <th data-field="units">Doporuč. ks</th>
            <th data-field="cost">Cena pozice</th>
            <th data-field="profit" class="sorted">Zisk pozice</th>
          </tr>
        </thead>
        <tbody id="table-body"></tbody>
      </table>
    </div>
  </section>

  <footer class="notes">
    <strong>Metodika.</strong> Ceny = top-of-book, tedy nejlepší stojící nákupní/prodejní order na stanici Jita 4-4 — stejné číslo jako „Market Buy/Sell" v klientu. Marže a zisk/kus jsou po odečtení broker fee (1,382 %) a sales tax (3,375 %). Doporučené množství = 15 % 14denního průměrného denního obratu — čistě likviditní úvaha, není vázáno na žádný kapitálový strop. Podrobná historie metodiky a známé problémy: viz docs/eve-jita-scanner-ops.md v repu.
  </footer>

</div>

<div class="viz-tooltip" id="tooltip"></div>

<script>
const RAW_DATA = __DATA_JSON__;

const COLOR_HEX = (() => {
  const s = getComputedStyle(document.documentElement);
  return {
    low: s.getPropertyValue('--slot-1').trim(),
    mid: s.getPropertyValue('--slot-2').trim(),
    high: s.getPropertyValue('--slot-3').trim(),
    accent: s.getPropertyValue('--accent').trim(),
  };
})();

function tier(buy) {
  if (buy < 10000) return 'low';
  if (buy < 1000000) return 'mid';
  return 'high';
}
RAW_DATA.forEach(d => { d.tier = tier(d.buy); });

function fmtISK(v, compact = true) {
  const sign = v < 0 ? '-' : '';
  v = Math.abs(v);
  if (!compact) return sign + Math.round(v).toLocaleString('cs-CZ') + ' ISK';
  if (v >= 1e9) return sign + (v / 1e9).toFixed(2) + ' mld';
  if (v >= 1e6) return sign + (v / 1e6).toFixed(1) + ' mil';
  if (v >= 1e3) return sign + (v / 1e3).toFixed(1) + ' tis';
  return sign + Math.round(v).toString();
}
function fmtNum(v) {
  if (v >= 1e6) return (v / 1e6).toFixed(1) + 'M';
  if (v >= 1e3) return (v / 1e3).toFixed(1) + 'k';
  return Math.round(v).toLocaleString('cs-CZ');
}
function fmtPct(v) { return v.toFixed(1) + '%'; }

const tooltip = document.getElementById('tooltip');
function showTooltip(evt, html) {
  tooltip.innerHTML = html;
  tooltip.classList.add('show');
  positionTooltip(evt);
}
function positionTooltip(evt) {
  const pad = 14;
  let x = evt.clientX + pad;
  let y = evt.clientY + pad;
  const tw = 240, th = 120;
  if (x + tw > window.innerWidth) x = evt.clientX - tw - pad;
  if (y + th > window.innerHeight) y = evt.clientY - th - pad;
  tooltip.style.left = x + 'px';
  tooltip.style.top = y + 'px';
}
function hideTooltip() { tooltip.classList.remove('show'); }

const state = { search: '', minMargin: 0, minVolume: 0, tiers: { low: true, mid: true, high: true }, sortField: 'profit', sortDir: 'desc' };

function getFiltered() {
  const q = state.search.trim().toLowerCase();
  return RAW_DATA.filter(d => {
    if (q && !d.name.toLowerCase().includes(q)) return false;
    if (d.margin < state.minMargin) return false;
    if (d.volume < state.minVolume) return false;
    if (!state.tiers[d.tier]) return false;
    return true;
  });
}

function renderKPIs(filtered) {
  const el = document.getElementById('kpi-row');
  const top20 = [...filtered].sort((a, b) => b.profit - a.profit).slice(0, 20);
  const top20Profit = top20.reduce((s, d) => s + d.profit, 0);
  const avgMargin = filtered.length ? filtered.reduce((s, d) => s + d.margin, 0) / filtered.length : 0;
  const topItem = top20[0];
  el.innerHTML = `
    <div class="kpi-card">
      <div class="kpi-label">Kandidátů ve filtru</div>
      <div class="kpi-value num">${filtered.length}</div>
      <div class="kpi-sub">z celkem ${RAW_DATA.length} položek</div>
    </div>
    <div class="kpi-card accent">
      <div class="kpi-label">Potenciál top 20 (zisk)</div>
      <div class="kpi-value num">${fmtISK(top20Profit)}</div>
      <div class="kpi-sub">ISK, kdyby šly všechny na trh</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Průměrná marže</div>
      <div class="kpi-value num">${avgMargin.toFixed(1)}%</div>
      <div class="kpi-sub">napříč filtrovanou sadou</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Nejlepší pozice</div>
      <div class="kpi-value" style="font-size:18px;">${topItem ? topItem.name : '—'}</div>
      <div class="kpi-sub">${topItem ? fmtISK(topItem.profit) + ' ISK zisk' : ''}</div>
    </div>
  `;
}

function renderBarChart(filtered) {
  const svg = document.getElementById('bar-chart');
  const W = 720, H = 420;
  const items = [...filtered].sort((a, b) => b.profit - a.profit).slice(0, 15);
  const marginL = 150, marginR = 70, marginT = 10, marginB = 10;
  const plotW = W - marginL - marginR;
  const rowH = items.length ? (H - marginT - marginB) / items.length : 0;
  const barH = Math.min(22, rowH * 0.62);
  const maxVal = items.length ? Math.max(...items.map(d => d.profit)) : 1;
  let s = '';
  items.forEach((d, i) => {
    const y = marginT + i * rowH + (rowH - barH) / 2;
    const w = Math.max(2, (d.profit / maxVal) * plotW);
    s += `<g class="bar-g" data-idx="${i}">
      <text x="${marginL - 10}" y="${y + barH / 2 + 4}" text-anchor="end" class="bar-label">${escapeHtml(truncate(d.name, 20))}</text>
      <rect x="${marginL}" y="${y}" width="${w}" height="${barH}" rx="4" fill="${COLOR_HEX[d.tier]}" opacity="0.92" style="cursor:pointer"></rect>
      <text x="${marginL + w + 8}" y="${y + barH / 2 + 4}" class="bar-value">${fmtISK(d.profit)}</text>
    </g>`;
  });
  svg.innerHTML = `<line class="axis-line" x1="${marginL}" y1="${marginT}" x2="${marginL}" y2="${H - marginB}"></line>${s}`;
  svg.querySelectorAll('.bar-g').forEach((g, i) => {
    const d = items[i];
    const rect = g.querySelector('rect');
    rect.addEventListener('mousemove', (evt) => showTooltip(evt, tooltipHtml(d)));
    rect.addEventListener('mouseleave', hideTooltip);
  });
}

function renderScatter(filtered) {
  const svg = document.getElementById('scatter-chart');
  const W = 520, H = 420;
  const marginL = 52, marginR = 16, marginT = 16, marginB = 40;
  const plotW = W - marginL - marginR;
  const plotH = H - marginT - marginB;
  if (!filtered.length) { svg.innerHTML = ''; return; }
  const xMax = 100, xMin = 0;
  const volVals = filtered.map(d => Math.max(d.volume, 1));
  const yMin = Math.max(10, Math.min(...volVals) * 0.7);
  const yMax = Math.max(...volVals) * 1.4;
  const logMin = Math.log10(yMin), logMax = Math.log10(yMax);
  const profitVals = filtered.map(d => d.profit);
  const pMax = Math.max(...profitVals), pMin = Math.min(...profitVals);
  const rScale = (p) => { const t = pMax > pMin ? (p - pMin) / (pMax - pMin) : 0.5; return 3 + Math.sqrt(t) * 15; };
  const xPix = (m) => marginL + ((m - xMin) / (xMax - xMin)) * plotW;
  const yPix = (v) => marginT + plotH - ((Math.log10(Math.max(v, 1)) - logMin) / (logMax - logMin)) * plotH;
  let grid = '';
  const decades = [10, 100, 1000, 10000, 100000, 1000000, 10000000, 100000000, 1000000000];
  decades.forEach(dv => {
    if (dv < yMin || dv > yMax) return;
    const y = yPix(dv);
    grid += `<line class="gridline" x1="${marginL}" y1="${y}" x2="${W - marginR}" y2="${y}"></line>`;
    grid += `<text x="${marginL - 8}" y="${y + 3}" text-anchor="end" font-size="10">${fmtNum(dv)}</text>`;
  });
  [0, 20, 40, 60, 80, 100].forEach(mv => {
    const x = xPix(mv);
    grid += `<text x="${x}" y="${H - marginB + 16}" text-anchor="middle" font-size="10">${mv}%</text>`;
  });
  grid += `<line class="axis-line" x1="${marginL}" y1="${marginT}" x2="${marginL}" y2="${H - marginB}"></line>`;
  grid += `<line class="axis-line" x1="${marginL}" y1="${H - marginB}" x2="${W - marginR}" y2="${H - marginB}"></line>`;
  let pts = '';
  filtered.forEach((d, i) => {
    const cx = xPix(Math.min(d.margin, xMax));
    const cy = yPix(d.volume);
    const r = rScale(d.profit);
    pts += `<circle class="scatter-pt" data-idx="${i}" cx="${cx}" cy="${cy}" r="${r}" fill="${COLOR_HEX[d.tier]}" fill-opacity="0.55" stroke="${COLOR_HEX[d.tier]}" stroke-width="1.2" style="cursor:pointer"></circle>`;
  });
  svg.innerHTML = grid + pts;
  svg.querySelectorAll('.scatter-pt').forEach((c) => {
    const i = +c.getAttribute('data-idx');
    const d = filtered[i];
    c.addEventListener('mousemove', (evt) => showTooltip(evt, tooltipHtml(d)));
    c.addEventListener('mouseleave', hideTooltip);
  });
}

function tooltipHtml(d) {
  return `<div class="tt-title">${escapeHtml(d.name)}</div>
    <div class="tt-row"><span>Marže</span><span>${fmtPct(d.margin)}</span></div>
    <div class="tt-row"><span>Obrat/den</span><span>${fmtNum(d.volume)} ks</span></div>
    <div class="tt-row"><span>Nákup / Prodej</span><span>${fmtISK(d.buy)} / ${fmtISK(d.sell)}</span></div>
    <div class="tt-row"><span>Doporuč. ks</span><span>${fmtNum(d.units)}</span></div>
    <div class="tt-row"><span>Zisk pozice</span><span>${fmtISK(d.profit)}</span></div>`;
}

function truncate(s, n) { return s.length > n ? s.slice(0, n - 1) + '…' : s; }
function escapeHtml(s) { return s.replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }

function renderTable(filtered) {
  const tbody = document.getElementById('table-body');
  document.getElementById('table-count').textContent = filtered.length + ' / ' + RAW_DATA.length + ' položek';
  const sorted = [...filtered].sort((a, b) => {
    const f = state.sortField;
    const cmp = a[f] < b[f] ? -1 : a[f] > b[f] ? 1 : 0;
    return state.sortDir === 'asc' ? cmp : -cmp;
  });
  if (!sorted.length) {
    tbody.innerHTML = `<tr class="empty-row"><td colspan="10">Žádná položka neodpovídá filtru.</td></tr>`;
    return;
  }
  const rows = sorted.slice(0, 200).map(d => `
    <tr>
      <td class="rank-cell">${d.rank}</td>
      <td><span class="tier-dot" style="background:${COLOR_HEX[d.tier]}"></span>${escapeHtml(d.name)}</td>
      <td class="num">${fmtISK(d.buy, true)}</td>
      <td class="num">${fmtISK(d.sell, true)}</td>
      <td class="num">${fmtPct(d.margin)}</td>
      <td class="num">${fmtISK(d.ppu, true)}</td>
      <td class="num">${fmtNum(d.volume)}</td>
      <td class="num">${fmtNum(d.units)}</td>
      <td class="num">${fmtISK(d.cost)}</td>
      <td class="num">${fmtISK(d.profit)}</td>
    </tr>`).join('');
  tbody.innerHTML = rows;
  document.querySelectorAll('.data-table thead th').forEach(th => th.classList.toggle('sorted', th.dataset.field === state.sortField));
}

document.querySelectorAll('.data-table thead th').forEach(th => {
  th.addEventListener('click', () => {
    const f = th.dataset.field;
    if (state.sortField === f) { state.sortDir = state.sortDir === 'asc' ? 'desc' : 'asc'; }
    else { state.sortField = f; state.sortDir = ['name'].includes(f) ? 'asc' : 'desc'; }
    render();
  });
});

document.getElementById('f-search').addEventListener('input', (e) => { state.search = e.target.value; render(); });
document.getElementById('f-margin').addEventListener('input', (e) => { state.minMargin = +e.target.value || 0; render(); });
document.getElementById('f-volume').addEventListener('input', (e) => { state.minVolume = +e.target.value || 0; render(); });
document.querySelectorAll('#tier-chips .chip').forEach(chip => {
  chip.addEventListener('click', () => {
    const t = chip.dataset.tier;
    state.tiers[t] = !state.tiers[t];
    chip.dataset.active = state.tiers[t];
    render();
  });
});
document.getElementById('reset-btn').addEventListener('click', () => {
  state.search = ''; state.minMargin = 0; state.minVolume = 0;
  state.tiers = { low: true, mid: true, high: true };
  document.getElementById('f-search').value = '';
  document.getElementById('f-margin').value = 0;
  document.getElementById('f-volume').value = 0;
  document.querySelectorAll('#tier-chips .chip').forEach(c => c.dataset.active = 'true');
  render();
});

function render() {
  const filtered = getFiltered();
  document.getElementById('filter-count').textContent = filtered.length + ' výsledků';
  renderKPIs(filtered);
  renderBarChart(filtered);
  renderScatter(filtered);
  renderTable(filtered);
}
render();
</script>
"""


def build_dashboard_html(rows, scan_date, out_path):
    html = DASHBOARD_TEMPLATE.replace("__DATA_JSON__", json.dumps(rows, ensure_ascii=False))
    html = html.replace("__SCAN_DATE__", scan_date)
    Path(out_path).write_text(html, encoding="utf-8")
    return out_path


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------
def build_xlsx_report(rows, scan_date, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ARIAL = "Arial"
    ISK_FMT = '#,##0" ISK"'
    PCT_FMT = '0.00"%"'
    UNIT_FMT = '#,##0'

    wb = Workbook()
    ws = wb.active
    ws.title = "Denní report"

    headers = [
        "Pořadí", "Položka", "Nákupní cena (ISK)", "Prodejní cena (ISK)", "Marže (%)",
        "Zisk/kus (ISK)", "Denní objem (14d prům.)", "Doporučené množství",
        "Cena pozice (ISK)", "Zisk pozice (ISK)", "Kumulativní cena (ISK)",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=ARIAL, bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name=ARIAL)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["rank"]).font = body_font
        ws.cell(row=i, column=2, value=r["name"]).font = body_font
        c = ws.cell(row=i, column=3, value=r["buy"]); c.font = body_font; c.number_format = ISK_FMT
        c = ws.cell(row=i, column=4, value=r["sell"]); c.font = body_font; c.number_format = ISK_FMT
        c = ws.cell(row=i, column=5, value=r["margin"]); c.font = body_font; c.number_format = PCT_FMT
        c = ws.cell(row=i, column=6, value=r["ppu"]); c.font = body_font; c.number_format = ISK_FMT
        c = ws.cell(row=i, column=7, value=r["volume"]); c.font = body_font; c.number_format = UNIT_FMT
        c = ws.cell(row=i, column=8, value=r["units"]); c.font = body_font; c.number_format = UNIT_FMT
        c = ws.cell(row=i, column=9, value=r["cost"]); c.font = body_font; c.number_format = ISK_FMT
        c = ws.cell(row=i, column=10, value=r["profit"]); c.font = body_font; c.number_format = ISK_FMT
        formula = "=I2" if i == 2 else f"=K{i-1}+I{i}"
        c = ws.cell(row=i, column=11, value=formula); c.font = body_font; c.number_format = ISK_FMT

    last_row = len(rows) + 1
    for idx, w in enumerate([8, 34, 16, 16, 10, 14, 18, 16, 16, 16, 18], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    notes = wb.create_sheet("Poznámky")
    notes_data = [
        ("Zdroj dat", f"eve-jita-poller (GCP Cloud Run Job), scan {scan_date}, region The Forge (Jita), ITEM_MODE=top_volume, TOP_N_ITEMS=750"),
        ("Filtrováno na", "min. 3 buy a 3 sell ordery na stanici Jita 4-4, min. 50 ks/den obratu, marže 0,1-100 % (strop jako pojistka proti zbytkovým výstřelkům u extrémně tenkých knih)"),
        ("Ceny", "top-of-book - nejlepší stojící nákupní/prodejní order na stanici Jita 4-4 (stejné číslo jako \"Market Buy/Sell\" v klientu). Historie metodiky (proč ne vážený průměr, proč ne min. počet orderů): viz docs/eve-jita-scanner-ops.md v repu."),
        ("Doporučené množství", "15 % průměrného denního objemu (14denní historie) - čistě otázka likvidity/obrátkovosti, NENÍ vázáno na žádný kapitálový strop"),
        ("Kumulativní cena", "běžící součet ceny pozic odshora - najdi řádek, kde součet překročí tvůj dnešní volný kapitál, a tam report pro dnešek zastav"),
        ("Marže/zisk", "po odečtení broker fee (1,382 %) a sales tax (3,375 %) - tvé reálné sazby"),
    ]
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 110
    for i, (label, text) in enumerate(notes_data, start=1):
        a = notes.cell(row=i, column=1, value=label)
        a.font = Font(name=ARIAL, bold=True)
        a.alignment = Alignment(vertical="top")
        b = notes.cell(row=i, column=2, value=text)
        b.font = Font(name=ARIAL)
        b.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    csv_path = Path(sys.argv[1])
    scan_date = sys.argv[2] if len(sys.argv) > 2 else "unknown-date"
    out_dir = csv_path.parent

    rows = load_rows(csv_path)
    print(f"Loaded {len(rows)} candidates from {csv_path}")

    xlsx_path = out_dir / f"jita_denni_report_{scan_date}.xlsx"
    build_xlsx_report(rows, scan_date, xlsx_path)
    print(f"Wrote {xlsx_path} (formulas NOT recalculated yet - run the xlsx skill's recalc.py)")

    html_path = out_dir / "jita_dashboard.html"
    build_dashboard_html(rows, scan_date, html_path)
    print(f"Wrote {html_path} (publish/update via the Artifact tool)")


if __name__ == "__main__":
    main()
